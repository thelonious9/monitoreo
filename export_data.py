import pandas as pd
import json
import os

def normalize_text(text):
    if not isinstance(text, str) or text == "N/A":
        return "N/A"
    # Basic trim and Title Case
    text = text.strip().capitalize()
    
    # Specific mapping for known variations/typos/acronyms
    mapping = {
        'Segobh': 'Segobh', # Keep as is or map if needed
        'Ssph': 'SSPH',
        'Difh': 'DIFH',
        'Ieeh': 'IEEH',
        'Teeh': 'TEEH',
        'Cdheh': 'CDHEH',
        'Oficiala mayor': 'Oficialía Mayor',
        'Oficiala_mayor': 'Oficialía Mayor',
        'Pgjeh': 'PGJEH',
        'Tsjeh': 'TSJEH',
        'Uaeh': 'UAEH',
        'Seph': 'SEPH',
        'Rrss': 'RRSS',
        'Tv local': 'TV local',
        'Tv nacional': 'TV nacional',
        'Tv': 'TV',
        'Portales locales': 'Portales Locales'
    }
    # Check if capitalized version is in mapping, otherwise return capitalized
    return mapping.get(text, text)

def export_to_json(excel_path, json_path):
    print(f"Reading {excel_path}...")
    try:
        df = pd.read_excel(excel_path, sheet_name='Fuente')
        
        # Mapping to fix encoding artifacts by using column positions
        print(f"Original columns: {df.columns.tolist()}")
        # We enforce the expected names directly for the first 12 columns
        target_cols = ['Fecha', 'Medio', 'Dependencia', 'Organismo', 'Género', 'Canal', 'Tema', 'Estatus', 'Municipio', 'Región', 'Clasificación', 'Agrupación']
        if len(df.columns) >= 12:
            current_cols = list(df.columns)
            for i in range(len(target_cols)):
                current_cols[i] = target_cols[i]
            df.columns = current_cols
        print(f"Final columns: {df.columns.tolist()}")

        # Extra check: row count with openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            ws = wb['Fuente']
            print(f"openpyxl Row Count: {ws.max_row}")
        except:
            pass
        
        # Normalization of fields
        fields_to_normalize = ["Dependencia", "Tema", "Estatus", "Municipio", "Región", "Canal", "Medio", "Género"]
        for col in fields_to_normalize:
            if col in df.columns:
                df[col] = df[col].apply(normalize_text)
        
        # Convert Timestamps to strings
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%d/%m/%Y')
        
        # FIX: Correct January 2025 dates to 2026 as per user request
        if 'Fecha' in df.columns:
            df['Fecha'] = df['Fecha'].apply(lambda x: x.replace('/01/2025', '/01/2026') if isinstance(x, str) else x)
        
        # Fill NaN with "N/A"
        df = df.fillna("N/A")
        
        # Ensure critical columns exist for the dashboard
        required_cols = ["Dependencia", "Tema", "Estatus", "Medio", "Municipio", "Región", "Canal", "Clasificación", "Agrupación", "Cobertura", "Nivel", "Género"]
        for col in required_cols:
            if col not in df.columns:
                df[col] = "N/A"
        
        data = df.to_dict(orient='records')
        
        print(f"Writing {len(data)} records to {json_path}...")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
        print("Success!")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    base_dir = r"c:\Users\vivie\Desktop\Dashboard"
    export_to_json(os.path.join(base_dir, "Data", "Fuente.xlsm"), os.path.join(base_dir, "Data", "datos.json"))
